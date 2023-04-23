import streamlit as st
import openpyxl
import pickle
import pandas as pd
import matplotlib.pyplot as plt
from PIL import Image

# Create the Streamlit app
image = Image.open('logo.png')
#import streamlit as st
from PIL import Image

# Load image from file
image = Image.open("logo.png")

# Create a layout with two columns
col1, col2 = st.columns([1, 3])

# Add the image to the first column
with col1:
    st.image(image)

# Add the title to the second column
with col2:
    st.title('Personality Prediction App')

subheader = st.markdown("Your personality results will be generated below:")
st.sidebar.title("Test your personality here!")

# Define the list of questions
questions = [
    "I am the life of the party.",
    "I dont talk a lot.",
    "I feel comfortable around people.",
    "I keep in the background.",
    "I start conversations.",
    "I have little to say.",
    "I talk to a lot of different people at parties.",
    "I dont like to draw attention to myself.",
    "I dont mind being the center of attention.",
    "I am quiet around strangers.",
    "I get stressed out easily.",
    "I am relaxed most of the time.",
    "I worry about things.",
    "I seldom feel blue.",
    "I am easily disturbed.",
    "I get upset easily.",
    "I change my mood a lot.",
    "I have frequent mood swings.",
    "I get irritated easily.",
    "I often feel blue.",
    "I feel little concern for others.",
    "I am interested in people.",
    "I insult people.",
    "I sympathize with others feelings.",
    "I am not interested in other peoples problems.",
    "I have a soft heart.",
    "I am not really interested in others.",
    "I take time out for others.",
    "I feel others emotions.",
    "I make people feel at ease.",
    "I am always prepared.",
    "I leave my belongings around.",
    "I pay attention to details.",
    "I make a mess of things.",
    "I get chores done right away.",
    "I often forget to put things back in their proper place.",
    "I like order.",
    "I shirk my duties.",
    "I follow a schedule.",
    "I am exacting in my work.",
    "I have a rich vocabulary.",
    "I have difficulty understanding abstract ideas.",
    "I have a vivid imagination.",
    "I am not interested in abstract ideas.",
    "I have excellent ideas.",
    "I do not have a good imagination.",
    "I am quick to understand things.",
    "I use difficult words.",
    "I spend time reflecting on things.",
    "I am full of ideas."
]

# Create an empty list to store the user's responses
responses = []

# Create a dictionary to map the words to numbers
word_to_number = {"Highly Disagree": 1, "Disagree": 2, "Neutral": 3, "Agree": 4, "Highly Agree": 5}

# Display the questions in the sidebar
for i, question in enumerate(questions):
    response = st.sidebar.radio(question, ["Highly Disagree", "Disagree", "Neutral", "Agree", "Highly Agree"], key=i)
    responses.append(word_to_number[response])

# Create a function to generate the Excel file
def generate_excel_file():
    # Create a new workbook object
    wb = openpyxl.load_workbook('personality_result.xlsx')
    
    # Get the active sheet
    ws = wb.active
    
    # Write the questions in the first row
    # for i, question in enumerate(questions):
    #     ws.cell(row=1, column=i+1).value = question

    # Write the questions in the first row
    # Iterate over the cells in the first row and replace their values with new values
    new_values = ['EXT1','EXT2','EXT3','EXT4','EXT5','EXT6','EXT7','EXT8','EXT9','EXT10',
              'EST1','EST2','EST3','EST4','EST5','EST6','EST7','EST8','EST9','EST10',
              'AGR1','AGR2','AGR3','AGR4','AGR5','AGR6','AGR7','AGR8','AGR9','AGR10',
              'CSN1','CSN2','CSN3','CSN4','CSN5','CSN6','CSN7','CSN8','CSN9','CSN10',
              'OPN1','OPN2','OPN3','OPN4','OPN5','OPN6','OPN7','OPN8','OPN9','OPN10']

    for column in ws.iter_cols(min_row=1, max_row=1):
        for cell in column:
            if not cell.value:
                continue
            if cell.column > len(new_values):
                break
            cell.value = new_values[cell.column - 1]

    # Write the responses in the second row
    for i, response in enumerate(responses):
        ws.cell(row=2, column=i+1).value = response
    
    # Save the workbook
    wb.save('personality_result.xlsx')

# Create a function to reset the responses list
def reset_responses():
    responses.clear()

def generate_personality():
    with open('model.pkl','rb') as f:
        mp = pickle.load(f)

    user_data = pd.read_excel('personality_result.xlsx')
    user_personality = mp.predict(user_data)
    # st.table(user_data)

# Summing up the my question groups
    col_list = list(user_data)
    ext = col_list[0:10]
    est = col_list[10:20]
    agr = col_list[20:30]
    csn = col_list[30:40]
    opn = col_list[40:50]

    my_sums = pd.DataFrame()
    my_sums['extroversion'] = user_data[ext].sum(axis=1)/10
    my_sums['neurotic'] = user_data[est].sum(axis=1)/10
    my_sums['agreeable'] = user_data[agr].sum(axis=1)/10
    my_sums['conscientious'] = user_data[csn].sum(axis=1)/10
    my_sums['open'] = user_data[opn].sum(axis=1)/10
    my_sums['cluster'] = user_personality
    
    # Display the maximum value in my_sums
    # st.write('The maximum value in my_sums is:', my_sums.max().max())

    # Find the column name with the maximum value for each row
    max_cols = my_sums.idxmax(axis=1)
    st.write('Your most significant personality trait is', max_cols[0]," with a maximum score of", my_sums.max().max())
    
    # display different text depending on the most significant trait
    if max_cols[0] == 'open':
        st.write("This trait refers to a person's openness to new experiences, ideas, and ways of thinking. People high in openness tend to be imaginative, curious, and creative.")
    elif max_cols[0] == 'extr0version':
        st.write("This trait refers to a person's outgoingness, social energy, and assertiveness. People high in extraversion tend to be talkative, outgoing, and enjoy being around others.")
    elif max_cols[0] == 'agreeable':
            st.write("This trait refers to a person's tendency to be compassionate, cooperative, and empathetic. People high in agreeableness tend to be kind, helpful, and willing to put others first.")
    elif max_cols[0] == 'conscientious':
        st.write("This trait refers to a person's degree of organization and responsibility. People high in conscientiousness tend to be dependable, efficient, and goal-oriented.")
    elif max_cols[0] == 'neurotic':
        st.write("This trait refers to a person's emotional stability and tendency to experience negative emotions such as anxiety, anger, and sadness. People high in neuroticism tend to be more anxious, moody, and easily stressed.")

    st.subheader("Here are some suggested quotes relating to your personality!")

    if max_cols[0] == 'open':
        st.caption("_The mind is like a parachute. It doesn't work unless it's open._ - Frank Zappa")
        st.caption("_The best way to have a good idea is to have lots of ideas._ - Linus Pauling")
        st.caption("_The measure of intelligence is the ability to change._ - Albert Einstein")
        st.caption("_The greatest glory in living lies not in never falling, but in rising every time we fall._ - Nelson Mandela")
    elif max_cols[0] == 'extroversion':
        st.caption("I am not a product of my circumstances. I am a product of my decisions._ - Stephen Covey")       
        st.caption("Success is not final, failure is not fatal: it is the courage to continue that counts._ - Winston Churchill")
        st.caption("The best way to predict your future is to create it._ - Abraham Lincoln")
        st.caption("If you look at what you have in life, you'll always have more. If you look at what you don't have in life, you'll never have enough._ - Oprah Winfrey")
    elif max_cols[0] == 'agreeable':
        st.caption("We can never obtain peace in the outer world until we make peace with ourselves._ - Dalai Lama XIV (Agreeable people tend to prioritize harmony and avoid conflict.)")
        st.caption("The greatest glory in living lies not in never falling, but in rising every time we fall._ - Nelson Mandela (Agreeable people are often compassionate and empathetic.)")
        st.caption("In the end, we will remember not the words of our enemies, but the silence of our friends._ - Martin Luther King Jr. (Agreeable people value relationships and tend to avoid confrontations.)")
        st.caption("Happiness is not something ready-made. It comes from your own actions._ - Dalai Lama XIV (Agreeable people are often optimistic and have a positive outlook on life.)")
    elif max_cols[0] == 'conscientious':
        st.caption("Success is not final, failure is not fatal: it is the courage to continue that counts._ - Winston Churchill (known for his strong work ethic and attention to detail)")
        st.caption("The difference between ordinary and extraordinary is that little extra._ - Jimmy Johnson (former NFL coach known for his meticulous preparation and attention to detail)")
        st.caption("The successful warrior is the average man, with laser-like focus._ - Bruce Lee (known for his discipline and dedication to his craft)")
        st.caption("The only way to do great work is to love what you do._ - Steve Jobs (known for his attention to detail and drive for perfection)")
    elif max_cols[0] == 'neurotic':
        st.caption("Worrying is like a rocking chair, it gives you something to do, but it gets you nowhere._ - Glenn Turner")
        st.caption("Anxiety is a thin stream of fear trickling through the mind. If encouraged, it cuts a channel into which all other thoughts are drained._ - Arthur Somers Roche")
        st.caption("I have had a lot of worries in my life, most of which never happened._ - Mark Twain")
        st.caption("You must learn to let go. Release the stress. You were never in control anyway._ - Steve Maraboli")
        
    st.markdown("-------------------------------------------------------------------")
    st.markdown("-------------------------------------------------------------------")
    my_sum = my_sums.drop('cluster', axis=1)
    # create the bar plot
    fig, ax = plt.subplots()
    ax.bar(my_sum.columns, my_sum.iloc[0,:], color='green', alpha=0.2)
    ax.plot(my_sum.columns, my_sum.iloc[0,:], color='red')
    ax.set_title('Cluster')
    ax.set_xticklabels(my_sum.columns, rotation=45)
    ax.set_ylim(0,5.0)

# display the plot in Streamlit
    st.write("Bar Graph of your Persnality Trait representing the most significant")
    st.pyplot(fig)

    st.markdown("The distribution of your Personality Traits")
    st.table(my_sums)

# Add a "Submit" and "Reset" button
col1, col2 = st.sidebar.columns(2)
if col1.button("Submit"):
    subheader.empty()
    generate_excel_file()
    generate_personality()
    st.success("Responses saved to Excel file")
if col2.button("Reset"):
    reset_responses()
    st.success("Responses reset")